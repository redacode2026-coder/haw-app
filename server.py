import http.server
import socketserver
import json
import sqlite3
import os
import sys
import urllib.parse
import mimetypes
import base64
import uuid
import datetime
import csv
import io
import hashlib
import secrets
import threading
import time

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

PORT = int(os.environ.get("PORT", 8080))
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

PUBLIC_DIR = os.path.join(BASE_DIR, "public")
UPLOADS_DIR = os.path.join(PUBLIC_DIR, "uploads")
DB_PATH = os.path.join(BASE_DIR, "database.db")

os.makedirs(PUBLIC_DIR, exist_ok=True)
os.makedirs(UPLOADS_DIR, exist_ok=True)

# ----------------- Password Hashing -----------------
def hash_password(password, salt=None):
    if salt is None:
        salt = secrets.token_hex(16)
    hashed = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), 600000).hex()
    return hashed, salt

def verify_password(password, stored_hash, salt):
    import hmac as _hmac
    hashed = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), 600000).hex()
    return _hmac.compare_digest(hashed, stored_hash)

# ----------------- Session Store (in-memory) -----------------
_sessions = {}
_sessions_lock = threading.Lock()

def create_session(user_id, role, full_name, username):
    token = secrets.token_hex(32)
    expires = datetime.datetime.utcnow() + datetime.timedelta(hours=12)
    with _sessions_lock:
        _sessions[token] = {
            'user_id': user_id,
            'role': role,
            'full_name': full_name,
            'username': username,
            'expires': expires
        }
    return token

def get_session(token):
    if not token:
        return None
    with _sessions_lock:
        sess = _sessions.get(token)
        if not sess:
            return None
        if datetime.datetime.utcnow() > sess['expires']:
            del _sessions[token]
            return None
        return sess

def delete_session(token):
    with _sessions_lock:
        _sessions.pop(token, None)

# ----------------- Database Initialization -----------------
def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS members (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        membership_number TEXT,
        full_name TEXT NOT NULL,
        nickname TEXT,
        national_id TEXT NOT NULL,
        id_issued_by TEXT,
        id_issue_date TEXT,
        email TEXT,
        birth_date TEXT,
        address TEXT,
        governorate TEXT,
        electoral_district TEXT,
        syndicate TEXT,
        qualification TEXT,
        job_title TEXT,
        workplace TEXT,
        work_sector TEXT,
        phone TEXT,
        mobile TEXT,
        public_positions TEXT,
        activities TEXT,
        previous_parties_status TEXT,
        previous_parties_details TEXT,
        elections_nomination_status TEXT,
        elections_entities TEXT,
        elections_other_entity TEXT,
        elections_details TEXT,
        endorser_name TEXT,
        endorser_title TEXT,
        applicant_signature TEXT,
        photo_url TEXT,
        national_id_photo_url TEXT,
        membership_officer_opinion TEXT,
        membership_officer_name TEXT,
        status TEXT DEFAULT 'قيد المراجعة',
        notes TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )
    """)

    # Users table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        salt TEXT NOT NULL,
        full_name TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'viewer',
        permissions TEXT,
        is_active INTEGER NOT NULL DEFAULT 1,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )
    """)
    # Add permissions column if upgrading from older version
    try:
        cursor.execute("ALTER TABLE users ADD COLUMN permissions TEXT")
    except Exception:
        pass
    conn.commit()


    # Seed sample data if empty to show the dashboard immediately populated
    cursor.execute("SELECT COUNT(*) FROM members")
    count = cursor.fetchone()[0]
    if count == 0:
        sample_activities = json.dumps(["شبابي", "تنظيم", "علاقات عامة"], ensure_ascii=False)
        sample_parties = json.dumps(["حزب سابق (2018)", "", "", ""], ensure_ascii=False)
        sample_elections = json.dumps(["مراكز شباب"], ensure_ascii=False)
        
        cursor.execute("""
        INSERT INTO members (
            membership_number, full_name, nickname, national_id, id_issued_by, id_issue_date,
            email, birth_date, address, governorate, electoral_district, syndicate,
            qualification, job_title, workplace, work_sector, phone, mobile,
            public_positions, activities, previous_parties_status, previous_parties_details,
            elections_nomination_status, elections_entities, elections_other_entity, elections_details,
            endorser_name, endorser_title, status, created_at
        ) VALUES (
            'HW-BH-2026-001', 'أحمد محمود إبراهيم النجار', 'أحمد النجار', '29001011801234', 'سجل مدني دمنهور', '2022-05-15',
            'ahmed.elngar@example.com', '1990-01-01', 'دمنهور - شارع الجمهورية', 'البحيرة', 'دائرة دمنهور', 'نقابة المهندسين',
            'بكالوريوس هندسة مدنية', 'مهندس استشاري', 'شركة التعمير الحديثة', 'قطاع المقاولات والهندسة', '0453312345', '01012345678',
            'عضو مجلس إدارة مركز شباب دمنهور سابقاً', ?, 'yes', ?,
            'yes', ?, '', 'الترشح لانتخابات مجلس إدارة مركز الشباب 2021 والحصول على المركز الأول',
            'محمد عبد الفتاح القاضي', 'أمين التنظيم المساعد', 'معتمد', datetime('now', '-2 days')
        )
        """, (sample_activities, sample_parties, sample_elections))

        cursor.execute("""
        INSERT INTO members (
            membership_number, full_name, nickname, national_id, id_issued_by, id_issue_date,
            email, birth_date, address, governorate, electoral_district, syndicate,
            qualification, job_title, workplace, work_sector, phone, mobile,
            public_positions, activities, previous_parties_status, previous_parties_details,
            elections_nomination_status, elections_entities, elections_other_entity, elections_details,
            endorser_name, endorser_title, status, created_at
        ) VALUES (
            'HW-BH-2026-002', 'سارة كمال عبد العزيز الشريف', 'سارة الشريف', '29406151805678', 'سجل مدني إيتاي البارود', '2023-08-20',
            'sara.elsharif@example.com', '1994-06-15', 'إيتاي البارود - البحيرة', 'البحيرة', 'دائرة إيتاي البارود', 'نقابة المعلمين',
            'ماجستير مناهج وطرق تدريس', 'معلمة أولى لغة عربية', 'مدرسة الثورة الثانوية', 'قطاع التعليم العام', '', '01123456789',
            'رئيسة لجنة المرأة بجمعية تنمية المجتمع', ?, 'no', '["", "", "", ""]',
            'no', '[]', '', '',
            'د. شريف متولي', 'أمين لجنة التعليم', 'قيد المراجعة', datetime('now', '-1 hours')
        )
        """, (json.dumps(["مرأة", "ثقافي", "علمي"], ensure_ascii=False),))

        conn.commit()

    # Seed default users if none exist
    cursor.execute("SELECT COUNT(*) FROM users")
    if cursor.fetchone()[0] == 0:
        default_users = [
            ('admin',    'admin123',    'مدير النظام',              'admin'),
            ('editor',   'editor123',   'مسؤول التسجيل والاستمارات', 'editor'),
            ('reviewer', 'reviewer123', 'مشرف ومراجع الطلبات',       'reviewer'),
            ('viewer',   'viewer123',   'مستعرض البيانات',           'viewer'),
        ]
        for uname, pwd, fname, role in default_users:
            h, s = hash_password(pwd)
            cursor.execute(
                "INSERT INTO users (username, password_hash, salt, full_name, role) VALUES (?,?,?,?,?)",
                (uname, h, s, fname, role)
            )
        conn.commit()

    conn.close()

# --------------- Default permissions per role ---------------
DEFAULT_PERMISSIONS = {
    'admin':    ['view_members', 'view_stats', 'add_member', 'edit_member',
                 'delete_member', 'approve_member', 'export_data', 'print_card', 'manage_users'],
    'editor':   ['view_members', 'view_stats', 'add_member', 'edit_member',
                 'export_data', 'print_card'],
    'reviewer': ['view_members', 'view_stats', 'approve_member',
                 'export_data', 'print_card'],
    'viewer':   [],
}

def get_effective_permissions(role, custom_permissions_json):
    """Return effective permissions list: custom if set, else role defaults."""
    if custom_permissions_json:
        try:
            custom = json.loads(custom_permissions_json)
            if isinstance(custom, list):
                return custom
        except Exception:
            pass
    return list(DEFAULT_PERMISSIONS.get(role, []))

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# ----------------- Rate Limiting -----------------
_login_attempts = {}  # ip -> [(timestamp, ...)]
_login_lock = threading.Lock()
MAX_LOGIN_ATTEMPTS = 8
LOGIN_WINDOW = 300  # 5 minutes

def check_rate_limit(ip):
    now = datetime.datetime.utcnow()
    with _login_lock:
        attempts = _login_attempts.get(ip, [])
        # Remove old attempts outside window
        attempts = [t for t in attempts if (now - t).total_seconds() < LOGIN_WINDOW]
        _login_attempts[ip] = attempts
        if len(attempts) >= MAX_LOGIN_ATTEMPTS:
            return False
        attempts.append(now)
        return True

def _cleanup_rate_limit():
    """Background thread to clean up old rate limit entries."""
    while True:
        time.sleep(300)
        now = datetime.datetime.utcnow()
        with _login_lock:
            expired = [ip for ip, attempts in _login_attempts.items()
                       if not any((now - t).total_seconds() < LOGIN_WINDOW for t in attempts)]
            for ip in expired:
                del _login_attempts[ip]

# ----------------- HTTP Request Handler -----------------
class PartyAppHandler(http.server.SimpleHTTPRequestHandler):
    MAX_BODY_SIZE = 10 * 1024 * 1024  # 10MB

    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=PUBLIC_DIR, **kwargs)

    def _read_body(self):
        content_length = int(self.headers.get('Content-Length', 0))
        if content_length > self.MAX_BODY_SIZE:
            self._send_json({'error': 'حجم الطلب أكبر من المسموح'}, 413)
            return None
        return self.rfile.read(content_length).decode('utf-8')

    def end_headers(self):
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "DENY")
        self.send_header("Referrer-Policy", "strict-origin-when-cross-origin")
        super().end_headers()

    def _send_json(self, data, status_code=200):
        response_bytes = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status_code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(response_bytes)))
        origin = self.headers.get('Origin', '')
        allowed = ['https://web-production-10e79c.up.railway.app', 'http://localhost:8080', 'http://localhost:3000']
        self.send_header("Access-Control-Allow-Origin", origin if origin in allowed else allowed[0])
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
        self.wfile.write(response_bytes)

    def _get_session_from_request(self):
        """Extract session token from Authorization header or cookie."""
        auth = self.headers.get('Authorization', '')
        if auth.startswith('Bearer '):
            return get_session(auth[7:])
        cookie = self.headers.get('Cookie', '')
        for part in cookie.split(';'):
            part = part.strip()
            if part.startswith('session_token='):
                return get_session(part[14:])
        return None

    def _require_auth(self, roles=None):
        """Returns session dict if authorized, else sends 401/403 and returns None."""
        sess = self._get_session_from_request()
        if not sess:
            self._send_json({'success': False, 'error': 'يرجى تسجيل الدخول أولاً'}, 401)
            return None
        if roles and sess['role'] not in roles:
            self._send_json({'success': False, 'error': 'ليس لديك صلاحية لهذه العملية'}, 403)
            return None
        return sess

    def do_OPTIONS(self):
        self.send_response(200)
        origin = self.headers.get('Origin', '')
        allowed = ['https://web-production-10e79c.up.railway.app', 'http://localhost:8080', 'http://localhost:3000']
        self.send_header("Access-Control-Allow-Origin", origin if origin in allowed else allowed[0])
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization")
        self.end_headers()

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        query = urllib.parse.parse_qs(parsed.query)

        if path == "/api/auth/me":
            self.handle_auth_me()
        elif path == "/api/users":
            self.handle_get_users()
        elif path.startswith("/api/users/"):
            uid = path.split("/")[-1]
            self.handle_get_single_user(uid)
        elif path == "/api/members":
            self.handle_get_members(query)
        elif path.startswith("/api/members/"):
            member_id = path.split("/")[-1]
            self.handle_get_single_member(member_id)
        elif path == "/api/stats":
            self.handle_get_stats()
        elif path == "/api/export":
            self.handle_export_csv()
        else:
            super().do_GET()

    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path

        if path == "/api/auth/login":
            self.handle_auth_login()
        elif path == "/api/auth/logout":
            self.handle_auth_logout()
        elif path == "/api/users":
            self.handle_create_user()
        elif path == "/api/members":
            self.handle_create_member()
        elif path == "/api/upload":
            sess = self._require_permission('add_member')
            if sess:
                self.handle_upload_file()
        else:
            self._send_json({"error": "المسار غير موجود"}, 404)

    def do_PUT(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path

        if path.startswith("/api/users/"):
            uid = path.split("/")[-1]
            self.handle_update_user(uid)
        elif path.startswith("/api/members/"):
            member_id = path.split("/")[-1]
            self.handle_update_member(member_id)
        else:
            self._send_json({"error": "المسار غير موجود"}, 404)

    def do_DELETE(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path

        if path.startswith("/api/users/"):
            uid = path.split("/")[-1]
            self.handle_delete_user(uid)
        elif path.startswith("/api/members/"):
            member_id = path.split("/")[-1]
            self.handle_delete_member(member_id)
        else:
            self._send_json({"error": "المسار غير موجود"}, 404)

    # ----------------- API Handlers -----------------
    def handle_get_members(self, query):
        sess = self._require_permission('view_members')
        if not sess:
            return

        search = query.get("search", [""])[0].strip()
        activity = query.get("activity", [""])[0].strip()
        status = query.get("status", [""])[0].strip()
        governorate = query.get("governorate", [""])[0].strip()
        district = query.get("district", query.get("electoral_district", [""]))[0].strip()

        conn = get_db()
        cursor = conn.cursor()

        sql = "SELECT * FROM members WHERE 1=1"
        params = []

        if search:
            sql += " AND (full_name LIKE ? OR national_id LIKE ? OR membership_number LIKE ? OR mobile LIKE ? OR electoral_district LIKE ? OR job_title LIKE ?)"
            wildcard = f"%{search}%"
            params.extend([wildcard, wildcard, wildcard, wildcard, wildcard, wildcard])

        if activity:
            sql += " AND activities LIKE ?"
            params.append(f"%{activity}%")

        if status:
            sql += " AND status = ?"
            params.append(status)

        if governorate:
            sql += " AND governorate LIKE ?"
            params.append(f"%{governorate}%")

        if district:
            sql += " AND electoral_district LIKE ?"
            params.append(f"%{district}%")

        sql += " ORDER BY id DESC"
        cursor.execute(sql, params)
        rows = [dict(r) for r in cursor.fetchall()]
        conn.close()

        # Parse JSON fields
        for row in rows:
            try:
                row["activities"] = json.loads(row["activities"]) if row["activities"] else []
            except Exception:
                row["activities"] = []
            try:
                row["previous_parties_details"] = json.loads(row["previous_parties_details"]) if row["previous_parties_details"] else []
            except Exception:
                row["previous_parties_details"] = []
            try:
                row["elections_entities"] = json.loads(row["elections_entities"]) if row["elections_entities"] else []
            except Exception:
                row["elections_entities"] = []

        self._send_json({"success": True, "count": len(rows), "data": rows})

    def handle_get_single_member(self, member_id):
        sess = self._require_permission('view_members')
        if not sess:
            return

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM members WHERE id = ?", (member_id,))
        row = cursor.fetchone()
        conn.close()

        if not row:
            self._send_json({"success": False, "error": "العضو غير موجود"}, 404)
            return

        data = dict(row)
        try:
            data["activities"] = json.loads(data["activities"]) if data["activities"] else []
        except Exception:
            data["activities"] = []
        try:
            data["previous_parties_details"] = json.loads(data["previous_parties_details"]) if data["previous_parties_details"] else []
        except Exception:
            data["previous_parties_details"] = []
        try:
            data["elections_entities"] = json.loads(data["elections_entities"]) if data["elections_entities"] else []
        except Exception:
            data["elections_entities"] = []

        self._send_json({"success": True, "data": data})

    def _require_permission(self, perm):
        sess = self._get_session_from_request()
        if not sess:
            self._send_json({'success': False, 'error': 'يرجى تسجيل الدخول أولاً'}, 401)
            return None
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT role, permissions FROM users WHERE id=?', (sess['user_id'],))
        row = cursor.fetchone()
        conn.close()
        perms = get_effective_permissions(row['role'], row['permissions']) if row else []
        if perm not in perms:
            self._send_json({'success': False, 'error': 'ليس لديك صلاحية لهذه العملية'}, 403)
            return None
        return sess

    def handle_create_member(self):
        try:
            sess = self._require_permission('add_member')
            if not sess:
                return

            body = self._read_body()
            if body is None:
                return
            data = json.loads(body)

            full_name = data.get("full_name", "").strip()
            national_id = data.get("national_id", "").strip()

            if not full_name or not national_id:
                self._send_json({"success": False, "error": "الاسم رباعي والرقم القومي حقول إجبارية"}, 400)
                return

            # Validate national_id
            national_id = data.get("national_id", "").strip()
            if national_id and (not national_id.isdigit() or len(national_id) != 14):
                self._send_json({"success": False, "error": "الرقم القومي يجب أن يكون 14 رقم"}, 400)
                return

            # Validate email format if provided
            email = data.get("email", "").strip()
            if email and "@" not in email:
                self._send_json({"success": False, "error": "صيغة البريد الإلكتروني غير صحيحة"}, 400)
                return

            # Validate mobile format if provided
            mobile = data.get("mobile", "").strip()
            if mobile and (not mobile.isdigit() or len(mobile) != 11):
                self._send_json({"success": False, "error": "رقم المحمول يجب أن يكون 11 رقم"}, 400)
                return

            # Whitelist status values
            VALID_STATUSES = ['قيد المراجعة', 'معتمد', 'مقبول', 'مرفوض']
            status = data.get('status', 'قيد المراجعة')
            if status not in VALID_STATUSES:
                status = 'قيد المراجعة'

            # Handle base64 photo save if provided
            photo_url = data.get("photo_url", "")
            if photo_url and photo_url.startswith("data:image/"):
                photo_url = self._save_base64_file(photo_url, "photo")

            national_id_photo_url = data.get("national_id_photo_url", "")
            if national_id_photo_url and national_id_photo_url.startswith("data:image/"):
                national_id_photo_url = self._save_base64_file(national_id_photo_url, "national_id")

            # Generate default membership code
            year = datetime.datetime.now().year
            code = f"HW-BH-{year}-{str(uuid.uuid4().int)[:5]}"

            conn = get_db()
            cursor = conn.cursor()
            cursor.execute("""
            INSERT INTO members (
                membership_number, full_name, nickname, national_id, id_issued_by, id_issue_date,
                email, birth_date, address, governorate, electoral_district, syndicate,
                qualification, job_title, workplace, work_sector, phone, mobile,
                public_positions, activities, previous_parties_status, previous_parties_details,
                elections_nomination_status, elections_entities, elections_other_entity, elections_details,
                endorser_name, endorser_title, applicant_signature, photo_url, national_id_photo_url,
                membership_officer_opinion, membership_officer_name, status, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get("membership_number", code),
                full_name,
                data.get("nickname", ""),
                national_id,
                data.get("id_issued_by", ""),
                data.get("id_issue_date", ""),
                data.get("email", ""),
                data.get("birth_date", ""),
                data.get("address", ""),
                data.get("governorate", "البحيرة"),
                data.get("electoral_district", ""),
                data.get("syndicate", ""),
                data.get("qualification", ""),
                data.get("job_title", ""),
                data.get("workplace", ""),
                data.get("work_sector", ""),
                data.get("phone", ""),
                data.get("mobile", ""),
                data.get("public_positions", ""),
                json.dumps(data.get("activities", []), ensure_ascii=False),
                data.get("previous_parties_status", "no"),
                json.dumps(data.get("previous_parties_details", ["", "", "", ""]), ensure_ascii=False),
                data.get("elections_nomination_status", "no"),
                json.dumps(data.get("elections_entities", []), ensure_ascii=False),
                data.get("elections_other_entity", ""),
                data.get("elections_details", ""),
                data.get("endorser_name", ""),
                data.get("endorser_title", ""),
                data.get("applicant_signature", ""),
                photo_url,
                national_id_photo_url,
                data.get("membership_officer_opinion", ""),
                data.get("membership_officer_name", ""),
                status,
                data.get("notes", "")
            ))
            new_id = cursor.lastrowid
            conn.commit()
            conn.close()

            self._send_json({
                "success": True,
                "message": "تم تسجيل استمارة العضوية بنجاح وحفظها في قاعدة البيانات",
                "id": new_id,
                "membership_number": data.get("membership_number", code)
            }, 201)
        except Exception as e:
            self._send_json({"success": False, "error": "حدث خطأ داخلي في الخادم"}, 500)

    def handle_update_member(self, member_id):
        try:
            sess = self._require_permission('edit_member')
            if not sess:
                return

            body = self._read_body()
            if body is None:
                return
            data = json.loads(body)

            # Validate national_id
            national_id = data.get("national_id", "").strip()
            if national_id and (not national_id.isdigit() or len(national_id) != 14):
                self._send_json({"success": False, "error": "الرقم القومي يجب أن يكون 14 رقم"}, 400)
                return

            # Validate email format if provided
            email = data.get("email", "").strip()
            if email and "@" not in email:
                self._send_json({"success": False, "error": "صيغة البريد الإلكتروني غير صحيحة"}, 400)
                return

            # Validate mobile format if provided
            mobile = data.get("mobile", "").strip()
            if mobile and (not mobile.isdigit() or len(mobile) != 11):
                self._send_json({"success": False, "error": "رقم المحمول يجب أن يكون 11 رقم"}, 400)
                return

            # Whitelist status values
            VALID_STATUSES = ['قيد المراجعة', 'معتمد', 'مقبول', 'مرفوض']
            status = data.get('status', 'قيد المراجعة')
            if status not in VALID_STATUSES:
                status = 'قيد المراجعة'

            conn = get_db()
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM members WHERE id = ?", (member_id,))
            existing = cursor.fetchone()
            if not existing:
                conn.close()
                self._send_json({"success": False, "error": "العضو غير موجود"}, 404)
                return

            photo_url = data.get("photo_url", existing["photo_url"])
            if photo_url and photo_url.startswith("data:image/"):
                photo_url = self._save_base64_file(photo_url, "photo")

            national_id_photo_url = data.get("national_id_photo_url", existing["national_id_photo_url"])
            if national_id_photo_url and national_id_photo_url.startswith("data:image/"):
                national_id_photo_url = self._save_base64_file(national_id_photo_url, "national_id")

            cursor.execute("""
            UPDATE members SET
                membership_number = ?, full_name = ?, nickname = ?, national_id = ?, id_issued_by = ?, id_issue_date = ?,
                email = ?, birth_date = ?, address = ?, governorate = ?, electoral_district = ?, syndicate = ?,
                qualification = ?, job_title = ?, workplace = ?, work_sector = ?, phone = ?, mobile = ?,
                public_positions = ?, activities = ?, previous_parties_status = ?, previous_parties_details = ?,
                elections_nomination_status = ?, elections_entities = ?, elections_other_entity = ?, elections_details = ?,
                endorser_name = ?, endorser_title = ?, applicant_signature = ?, photo_url = ?, national_id_photo_url = ?,
                membership_officer_opinion = ?, membership_officer_name = ?, status = ?, notes = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """, (
                data.get("membership_number", existing["membership_number"]),
                data.get("full_name", existing["full_name"]),
                data.get("nickname", existing["nickname"]),
                data.get("national_id", existing["national_id"]),
                data.get("id_issued_by", existing["id_issued_by"]),
                data.get("id_issue_date", existing["id_issue_date"]),
                data.get("email", existing["email"]),
                data.get("birth_date", existing["birth_date"]),
                data.get("address", existing["address"]),
                data.get("governorate", existing["governorate"]),
                data.get("electoral_district", existing["electoral_district"]),
                data.get("syndicate", existing["syndicate"]),
                data.get("qualification", existing["qualification"]),
                data.get("job_title", existing["job_title"]),
                data.get("workplace", existing["workplace"]),
                data.get("work_sector", existing["work_sector"]),
                data.get("phone", existing["phone"]),
                data.get("mobile", existing["mobile"]),
                data.get("public_positions", existing["public_positions"]),
                json.dumps(data.get("activities", []), ensure_ascii=False) if isinstance(data.get("activities"), list) else data.get("activities", existing["activities"]),
                data.get("previous_parties_status", existing["previous_parties_status"]),
                json.dumps(data.get("previous_parties_details", []), ensure_ascii=False) if isinstance(data.get("previous_parties_details"), list) else data.get("previous_parties_details", existing["previous_parties_details"]),
                data.get("elections_nomination_status", existing["elections_nomination_status"]),
                json.dumps(data.get("elections_entities", []), ensure_ascii=False) if isinstance(data.get("elections_entities"), list) else data.get("elections_entities", existing["elections_entities"]),
                data.get("elections_other_entity", existing["elections_other_entity"]),
                data.get("elections_details", existing["elections_details"]),
                data.get("endorser_name", existing["endorser_name"]),
                data.get("endorser_title", existing["endorser_title"]),
                data.get("applicant_signature", existing["applicant_signature"]),
                photo_url,
                national_id_photo_url,
                data.get("membership_officer_opinion", existing["membership_officer_opinion"]),
                data.get("membership_officer_name", existing["membership_officer_name"]),
                status,
                data.get("notes", existing["notes"]),
                member_id
            ))
            conn.commit()
            conn.close()

            self._send_json({"success": True, "message": "تم تحديث بيانات العضو بنجاح"})
        except Exception as e:
            self._send_json({"success": False, "error": "حدث خطأ داخلي في الخادم"}, 500)

    def handle_delete_member(self, member_id):
        try:
            sess = self._require_permission('delete_member')
            if not sess:
                return

            conn = get_db()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM members WHERE id = ?", (member_id,))
            conn.commit()
            conn.close()
            self._send_json({"success": True, "message": "تم حذف سجل العضوية بنجاح"})
        except Exception as e:
            self._send_json({"success": False, "error": "حدث خطأ داخلي في الخادم"}, 500)

    def handle_get_stats(self):
        sess = self._require_permission('view_stats')
        if not sess:
            return

        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("SELECT COUNT(*) FROM members")
        total = cursor.fetchone()[0]

        cursor.execute("SELECT status, COUNT(*) FROM members GROUP BY status")
        status_counts = dict(cursor.fetchall())

        cursor.execute("SELECT governorate, COUNT(*) FROM members GROUP BY governorate")
        gov_counts = dict(cursor.fetchall())

        cursor.execute("SELECT activities FROM members")
        all_activities = cursor.fetchall()
        activity_stats = {}
        for row in all_activities:
            if row[0]:
                try:
                    acts = json.loads(row[0])
                    for act in acts:
                        activity_stats[act] = activity_stats.get(act, 0) + 1
                except Exception:
                    pass

        conn.close()
        self._send_json({
            "success": True,
            "data": {
                "total": total,
                "status_counts": status_counts,
                "governorates": gov_counts,
                "activities": activity_stats
            }
        })

    def handle_export_csv(self):
        sess = self._require_permission('export_data')
        if not sess:
            return

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM members ORDER BY id ASC")
        rows = cursor.fetchall()
        conn.close()

        if HAS_OPENPYXL:
            self._export_xlsx(rows)
        else:
            self._export_csv_fallback(rows)

    def _export_xlsx(self, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "سجل الأعضاء"

        headers = [
            "رقم المسلسل", "رقم العضوية", "الاسم رباعي", "اسم الشهرة", "الرقم القومي",
            "صادر من", "بتاريخ", "البريد الإلكتروني", "تاريخ الميلاد", "محل الإقامة",
            "المحافظة", "الدائرة الانتخابية", "النقابة", "المؤهل الدراسي", "الوظيفة",
            "محل العمل", "قطاع العمل", "الهاتف الأرضي", "المحمول", "المناصب العامة",
            "الأنشطة المشارك فيها", "انتماء حزبي سابق", "تفاصيل الأحزاب السابقة",
            "ترشح لانتخابات سابقة", "جهات الترشح", "بيانات النشاط السياسي",
            "اسم المؤيد", "صفة المؤيد", "الحالة", "تاريخ التسجيل"
        ]

        header_font = Font(name='Cairo', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        data_font = Font(name='Cairo', size=10)
        data_alignment = Alignment(vertical='center', wrap_text=True)

        gold_fill = PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        for row_idx, r in enumerate(rows, 2):
            r_dict = dict(r)
            try:
                acts = ", ".join(json.loads(r_dict["activities"])) if r_dict["activities"] else ""
            except Exception:
                acts = r_dict["activities"] or ""

            try:
                parties = ", ".join([p for p in json.loads(r_dict["previous_parties_details"]) if p]) if r_dict["previous_parties_details"] else ""
            except Exception:
                parties = r_dict["previous_parties_details"] or ""

            try:
                elections = ", ".join(json.loads(r_dict["elections_entities"])) if r_dict["elections_entities"] else ""
            except Exception:
                elections = r_dict["elections_entities"] or ""

            row_data = [
                r_dict["id"],
                r_dict["membership_number"] or "",
                r_dict["full_name"] or "",
                r_dict["nickname"] or "",
                r_dict["national_id"] or "",
                r_dict["id_issued_by"] or "",
                r_dict["id_issue_date"] or "",
                r_dict["email"] or "",
                r_dict["birth_date"] or "",
                r_dict["address"] or "",
                r_dict["governorate"] or "",
                r_dict["electoral_district"] or "",
                r_dict["syndicate"] or "",
                r_dict["qualification"] or "",
                r_dict["job_title"] or "",
                r_dict["workplace"] or "",
                r_dict["work_sector"] or "",
                r_dict["phone"] or "",
                r_dict["mobile"] or "",
                r_dict["public_positions"] or "",
                acts,
                "نعم" if r_dict["previous_parties_status"] == "yes" else "لا",
                parties,
                "نعم" if r_dict["elections_nomination_status"] == "yes" else "لا",
                elections,
                r_dict["elections_details"] or "",
                r_dict["endorser_name"] or "",
                r_dict["endorser_title"] or "",
                r_dict["status"] or "",
                r_dict["created_at"] or ""
            ]

            row_fill = gold_fill if row_idx % 2 == 0 else white_fill
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                cell.fill = row_fill

        col_widths = [8, 18, 28, 18, 18, 18, 14, 26, 14, 28, 14, 20, 18, 20, 22, 24, 22, 16, 14, 30, 24, 14, 30, 14, 24, 30, 22, 20, 14, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(rows) + 1}"
        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        xlsx_data = output.getvalue()

        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", 'attachment; filename="homat_alwatan_members.xlsx"')
        self.send_header("Content-Length", str(len(xlsx_data)))
        self.end_headers()
        self.wfile.write(xlsx_data)

    def _export_csv_fallback(self, rows):
        output = io.StringIO()
        output.write('\ufeff')
        writer = csv.writer(output)

        headers = [
            "رقم المسلسل", "رقم العضوية", "الاسم رباعي", "اسم الشهرة", "الرقم القومي",
            "صادر من", "بتاريخ", "البريد الإلكتروني", "تاريخ الميلاد", "محل الإقامة",
            "المحافظة", "الدائرة الانتخابية", "النقابة", "المؤهل الدراسي", "الوظيفة",
            "محل العمل", "قطاع العمل", "الهاتف الأرضي", "المحمول", "المناصب العامة",
            "الأنشطة المشارك فيها", "انتماء حزبي سابق", "تفاصيل الأحزاب السابقة",
            "ترشح لانتخابات سابقة", "جهات الترشح", "بيانات النشاط السياسي",
            "اسم المؤيد", "صفة المؤيد", "الحالة", "تاريخ التسجيل"
        ]
        writer.writerow(headers)

        for r in rows:
            r_dict = dict(r)
            try:
                acts = ", ".join(json.loads(r_dict["activities"])) if r_dict["activities"] else ""
            except Exception:
                acts = r_dict["activities"] or ""
            try:
                parties = ", ".join([p for p in json.loads(r_dict["previous_parties_details"]) if p]) if r_dict["previous_parties_details"] else ""
            except Exception:
                parties = r_dict["previous_parties_details"] or ""
            try:
                elections = ", ".join(json.loads(r_dict["elections_entities"])) if r_dict["elections_entities"] else ""
            except Exception:
                elections = r_dict["elections_entities"] or ""

            writer.writerow([
                r_dict["id"],
                r_dict["membership_number"] or "",
                r_dict["full_name"] or "",
                r_dict["nickname"] or "",
                r_dict["national_id"] or "",
                r_dict["id_issued_by"] or "",
                r_dict["id_issue_date"] or "",
                r_dict["email"] or "",
                r_dict["birth_date"] or "",
                r_dict["address"] or "",
                r_dict["governorate"] or "",
                r_dict["electoral_district"] or "",
                r_dict["syndicate"] or "",
                r_dict["qualification"] or "",
                r_dict["job_title"] or "",
                r_dict["workplace"] or "",
                r_dict["work_sector"] or "",
                r_dict["phone"] or "",
                r_dict["mobile"] or "",
                r_dict["public_positions"] or "",
                acts,
                "نعم" if r_dict["previous_parties_status"] == "yes" else "لا",
                parties,
                "نعم" if r_dict["elections_nomination_status"] == "yes" else "لا",
                elections,
                r_dict["elections_details"] or "",
                r_dict["endorser_name"] or "",
                r_dict["endorser_title"] or "",
                r_dict["status"] or "",
                r_dict["created_at"] or ""
            ])

        csv_data = output.getvalue().encode("utf-8-sig")
        self.send_response(200)
        self.send_header("Content-Type", "text/csv; charset=utf-8")
        self.send_header("Content-Disposition", 'attachment; filename="homat_alwatan_members.csv"')
        self.send_header("Content-Length", str(len(csv_data)))
        self.end_headers()
        self.wfile.write(csv_data)

    # ----------------- Auth Handlers -----------------
    def handle_auth_login(self):
        try:
            # Rate limiting check
            client_ip = self.client_address[0]
            if not check_rate_limit(client_ip):
                self._send_json({'success': False, 'error': 'تم تجاوز الحد المسموح لمحاولات تسجيل الدخول، يرجى المحاولة بعد 5 دقائق'}, 429)
                return

            body = self._read_body()
            if body is None:
                return
            data = json.loads(body)
            username = data.get('username', '').strip().lower()
            password = data.get('password', '')

            if not username or not password:
                self._send_json({'success': False, 'error': 'يرجى إدخال اسم المستخدم وكلمة المرور'}, 400)
                return

            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM users WHERE username=? AND is_active=1', (username,))
            row = cursor.fetchone()
            conn.close()

            if not row:
                self._send_json({'success': False, 'error': 'اسم المستخدم أو كلمة المرور غير صحيحة'}, 401)
                return

            user = dict(row)
            if not verify_password(password, user['password_hash'], user['salt']):
                self._send_json({'success': False, 'error': 'اسم المستخدم أو كلمة المرور غير صحيحة'}, 401)
                return

            token = create_session(user['id'], user['role'], user['full_name'], user['username'])
            self._send_json({
                'success': True,
                'token': token,
                'user': {
                    'id': user['id'],
                    'username': user['username'],
                    'full_name': user['full_name'],
                    'role': user['role'],
                    'permissions': get_effective_permissions(user['role'], user.get('permissions'))
                }
            })
        except Exception as e:
            self._send_json({'success': False, 'error': 'حدث خطأ داخلي في الخادم'}, 500)

    def handle_auth_logout(self):
        auth = self.headers.get('Authorization', '')
        if auth.startswith('Bearer '):
            delete_session(auth[7:])
        self._send_json({'success': True})

    def handle_auth_me(self):
        sess = self._get_session_from_request()
        if not sess:
            self._send_json({'success': False, 'error': 'غير مسجل الدخول'}, 401)
            return
        # Fetch fresh permissions from DB each time
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT role, permissions FROM users WHERE id=?', (sess['user_id'],))
        row = cursor.fetchone()
        conn.close()
        perms = get_effective_permissions(row['role'], row['permissions']) if row else []
        self._send_json({'success': True, 'user': {
            'id': sess['user_id'],
            'username': sess['username'],
            'full_name': sess['full_name'],
            'role': sess['role'],
            'permissions': perms
        }})

    # ----------------- Users CRUD -----------------
    def handle_get_users(self):
        sess = self._require_auth(roles=['admin'])
        if not sess:
            return
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT id, username, full_name, role, permissions, is_active, created_at FROM users ORDER BY id')
        rows = [dict(r) for r in cursor.fetchall()]
        conn.close()
        for u in rows:
            u['effective_permissions'] = get_effective_permissions(u['role'], u.get('permissions'))
            u['has_custom_permissions'] = bool(u.get('permissions'))
        self._send_json({'success': True, 'data': rows})

    def handle_get_single_user(self, uid):
        sess = self._require_auth(roles=['admin'])
        if not sess:
            return
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT id, username, full_name, role, permissions, is_active, created_at FROM users WHERE id=?', (uid,))
        row = cursor.fetchone()
        conn.close()
        if not row:
            self._send_json({'success': False, 'error': 'المستخدم غير موجود'}, 404)
            return
        u = dict(row)
        u['effective_permissions'] = get_effective_permissions(u['role'], u.get('permissions'))
        self._send_json({'success': True, 'data': u})


    def handle_create_user(self):
        sess = self._require_auth(roles=['admin'])
        if not sess:
            return
        try:
            body = self._read_body()
            if body is None:
                return
            data = json.loads(body)
            username  = data.get('username', '').strip().lower()
            password  = data.get('password', '').strip()
            full_name = data.get('full_name', '').strip()
            role      = data.get('role', 'viewer')
            custom_perms = data.get('permissions', None)

            if not username or not password or not full_name:
                self._send_json({'success': False, 'error': 'جميع الحقول مطلوبة'}, 400)
                return
            if role not in ('admin', 'editor', 'reviewer', 'viewer'):
                self._send_json({'success': False, 'error': 'الدور المحدد غير صالح'}, 400)
                return
            if password and len(password) < 6:
                self._send_json({'success': False, 'error': 'كلمة المرور يجب أن تكون 6 أحرف على الأقل'}, 400)
                return

            # Store custom permissions only when they differ from role defaults
            perms_json = None
            if isinstance(custom_perms, list):
                default = DEFAULT_PERMISSIONS.get(role, [])
                if sorted(custom_perms) != sorted(default):
                    perms_json = json.dumps(custom_perms, ensure_ascii=False)

            hashed, salt = hash_password(password)
            conn = get_db()
            cursor = conn.cursor()
            try:
                cursor.execute(
                    'INSERT INTO users (username, password_hash, salt, full_name, role, permissions) VALUES (?,?,?,?,?,?)',
                    (username, hashed, salt, full_name, role, perms_json)
                )
                conn.commit()
                new_id = cursor.lastrowid
            except sqlite3.IntegrityError:
                conn.close()
                self._send_json({'success': False, 'error': 'اسم المستخدم مستخدم بالفعل'}, 409)
                return
            conn.close()
            self._send_json({'success': True, 'id': new_id})
        except Exception as e:
            self._send_json({'success': False, 'error': 'حدث خطأ داخلي في الخادم'}, 500)

    def handle_update_user(self, uid):
        sess = self._require_auth(roles=['admin'])
        if not sess:
            return
        try:
            body = self._read_body()
            if body is None:
                return
            data = json.loads(body)
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM users WHERE id=?', (uid,))
            existing = cursor.fetchone()
            if not existing:
                conn.close()
                self._send_json({'success': False, 'error': 'المستخدم غير موجود'}, 404)
                return
            existing     = dict(existing)
            full_name    = data.get('full_name', existing['full_name']).strip()
            role         = data.get('role', existing['role'])
            is_active    = data.get('is_active', existing['is_active'])
            new_password = data.get('password', '').strip()
            custom_perms = data.get('permissions', None)

            if role not in ('admin', 'editor', 'reviewer', 'viewer'):
                conn.close()
                self._send_json({'success': False, 'error': 'الدور المحدد غير صالح'}, 400)
                return
            if new_password and len(new_password) < 6:
                conn.close()
                self._send_json({'success': False, 'error': 'كلمة المرور يجب أن تكون 6 أحرف على الأقل'}, 400)
                return

            # Build permissions JSON: None means "use role defaults"
            if isinstance(custom_perms, list):
                default = DEFAULT_PERMISSIONS.get(role, [])
                if sorted(custom_perms) != sorted(default):
                    perms_json = json.dumps(custom_perms, ensure_ascii=False)
                else:
                    perms_json = None   # reset to defaults (no override)
            else:
                perms_json = existing.get('permissions')  # keep existing value

            if new_password:
                hashed, salt = hash_password(new_password)
                cursor.execute(
                    'UPDATE users SET full_name=?, role=?, is_active=?, password_hash=?, salt=?, permissions=? WHERE id=?',
                    (full_name, role, int(is_active), hashed, salt, perms_json, uid)
                )
                # Revoke all sessions for this user
                with _sessions_lock:
                    to_delete = [k for k, v in _sessions.items() if v['user_id'] == int(uid)]
                    for k in to_delete:
                        del _sessions[k]
            else:
                cursor.execute(
                    'UPDATE users SET full_name=?, role=?, is_active=?, permissions=? WHERE id=?',
                    (full_name, role, int(is_active), perms_json, uid)
                )
            conn.commit()
            conn.close()
            self._send_json({'success': True})
        except Exception as e:
            self._send_json({'success': False, 'error': 'حدث خطأ داخلي في الخادم'}, 500)

    def handle_delete_user(self, uid):
        sess = self._require_auth(roles=['admin'])
        if not sess:
            return
        if str(uid) == str(sess['user_id']):
            self._send_json({'success': False, 'error': 'لا يمكنك حذف حسابك الخاص'}, 400)
            return
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users WHERE id=?', (uid,))
        existing = cursor.fetchone()
        if not existing:
            conn.close()
            self._send_json({'success': False, 'error': 'المستخدم غير موجود'}, 404)
            return
        existing = dict(existing)
        # Prevent deleting the last admin
        if existing.get('role') == 'admin':
            cursor2 = conn.cursor()
            cursor2.execute("SELECT COUNT(*) FROM users WHERE role='admin' AND is_active=1")
            admin_count = cursor2.fetchone()[0]
            if admin_count <= 1:
                conn.close()
                self._send_json({'success': False, 'error': 'لا يمكن حذف آخر مدير نظام'}, 400)
                return
        cursor.execute('DELETE FROM users WHERE id=?', (uid,))
        conn.commit()
        conn.close()
        self._send_json({'success': True})

    def _save_base64_file(self, data_uri, prefix):
        try:
            header, encoded = data_uri.split(",", 1)
            ext = "jpg"
            if "png" in header:
                ext = "png"
            elif "jpeg" in header:
                ext = "jpg"
            elif "webp" in header:
                ext = "webp"

            filename = f"{prefix}_{int(datetime.datetime.now().timestamp())}_{uuid.uuid4().hex[:6]}.{ext}"
            file_path = os.path.join(UPLOADS_DIR, filename)
            
            decoded = base64.b64decode(encoded)
            if len(decoded) > 5 * 1024 * 1024:  # 5MB max
                return ""
            with open(file_path, "wb") as f:
                f.write(decoded)
            
            return f"uploads/{filename}"
        except Exception as e:
            print(f"Error saving base64 file: {e}")
            return ""

def _cleanup_sessions_loop():
    while True:
        time.sleep(300)
        now = datetime.datetime.utcnow()
        with _sessions_lock:
            expired = [k for k, v in _sessions.items() if now > v['expires']]
            for k in expired:
                del _sessions[k]

def run_server():
    socketserver.TCPServer.allow_reuse_address = True
    
    # Auto open browser only in local dev (not on production server)
    is_production = bool(os.environ.get("RAILWAY_STATIC_URL") or os.environ.get("RENDER") or os.environ.get("DYNO") or os.environ.get("PORT"))
    if not is_production:
        def open_browser():
            import time
            import webbrowser
            time.sleep(1.2)
            try:
                webbrowser.open(f"http://localhost:{PORT}")
            except Exception:
                pass
        threading.Thread(target=open_browser, daemon=True).start()

    threading.Thread(target=_cleanup_sessions_loop, daemon=True).start()
    threading.Thread(target=_cleanup_rate_limit, daemon=True).start()

    with socketserver.ThreadingTCPServer(("0.0.0.0", PORT), PartyAppHandler) as httpd:
        print(f"==================================================")
        print(f"  حزب حماة الوطن - تطبيق إدارة العضوية وقاعدة البيانات")
        print(f"  الخادم يعمل الآن على الرابط: http://localhost:{PORT}")
        print(f"==================================================")
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\nتم إيقاف الخادم.")

if __name__ == "__main__":
    init_db()
    run_server()
